import pandas as pd
import os
from openpyxl import load_workbook


class GameResultsLogger:
    def __init__(self, filename="game_results.xlsx"):
        self.filename = filename
        self._current_game = None
        self._results = {
            'GameID': None,
            'Mines': 0,
            'BoardSize': None,
            'TotalMoves': 0,
            'NeighbourDeductionMoves': 0,
            'ClusterInferenceMoves': 0,
            'RandomForestMoves': 0,
            'MonteCarloMoves': 0,
            'Result': None,
            'LastAlgorithm': None
        }
        self._ensure_file_exists()

    def _ensure_file_exists(self):
        if not os.path.isfile(self.filename):
            columns = [
                'GameID', 'Mines', 'BoardSize', 'TotalMoves',
                'NeighbourDeductionMoves', 'ClusterInferenceMoves',
                'RandomForestMoves', 'MonteCarloMoves',
                'NeighbourDeductionPercentage', 'ClusterInferencePercentage',
                'RandomForestPercentage', 'MonteCarloPercentage', 'Result', 'LastAlgorithm',
                'Wins', 'Losses', 'NeighbourDeductionPercentage', 'ClusterInferencePercentage',
                'RandomForestPercentage', 'MonteCarloPercentage'
            ]
            df = pd.DataFrame(columns=columns)
            try:
                df.to_excel(self.filename, index=False, engine='openpyxl')
                print(f"Plik {self.filename} utworzony pomyślnie.")
            except Exception as e:
                print(f"Błąd podczas tworzenia pliku Excel: {e}")

    def start_new_game(self, game_id, rows, cols, mines):
        self._current_game = game_id
        self._results['GameID'] = game_id
        self._results['Mines'] = mines
        self._results['BoardSize'] = f"{rows}x{cols}"
        self._results['TotalMoves'] = 0
        self._results['NeighbourDeductionMoves'] = 0
        self._results['ClusterInferenceMoves'] = 0
        self._results['RandomForestMoves'] = 0
        self._results['MonteCarloMoves'] = 0

    def log_move(self, algorithm):
        self._results['TotalMoves'] += 1
        self._results['LastAlgorithm'] = algorithm
        if algorithm == "Neighbour Deduction":
            self._results['NeighbourDeductionMoves'] += 1
        elif algorithm == "Cluster Inference":
            self._results['ClusterInferenceMoves'] += 1
        elif algorithm == "Random Forest":
            self._results['RandomForestMoves'] += 1
        elif algorithm == "Monte Carlo":
            self._results['MonteCarloMoves'] += 1

    def finalize_results(self, result):
        self._results['Result'] = result
        if self._results['TotalMoves'] > 0:
            self._results['NeighbourDeductionPercentage'] = (self._results['NeighbourDeductionMoves'] / self._results['TotalMoves']) * 100
            self._results['ClusterInferencePercentage'] = (self._results['ClusterInferenceMoves'] / self._results['TotalMoves']) * 100
            self._results['RandomForestPercentage'] = (self._results['RandomForestMoves'] / self._results['TotalMoves']) * 100
            self._results['MonteCarloPercentage'] = (self._results['MonteCarloMoves'] / self._results['TotalMoves']) * 100

        self._save_results()

    def _save_results(self):
        try:
            if os.path.isfile(self.filename):
                existing_data = pd.read_excel(self.filename, engine='openpyxl')
            else:
                existing_data = pd.DataFrame()

            new_data = pd.DataFrame([self._results])

            updated_data = pd.concat([existing_data, new_data], ignore_index=True)

            updated_data.to_excel(self.filename, index=False, engine='openpyxl')

            self._add_excel_formulas()

            print(f"Wyniki zapisane do {self.filename}")
        except Exception as e:
            print(f"Błąd podczas zapisywania wyników do Excela: {e}")

    def _add_excel_formulas(self):
        try:
            wb = load_workbook(self.filename)
            ws = wb.active

            ws['O2'] = '=COUNTIF(M:M,"win")'
            ws['O3'] = '=O2/Q4'
            ws['P2'] = '=COUNTIF(M:M,"loss")'
            ws['P3'] = '=P2/Q4'
            ws['Q2'] = '=AVERAGE(I:I)'
            ws['Q3'] = 'Total Games'
            ws['Q4'] = '=O2+P2'
            ws['R2'] = '=AVERAGE(J:J)'
            ws['S2'] = '=AVERAGE(K:K)'
            ws['T2'] = '=AVERAGE(L:L)'
            ws['Q5'] = 'Neighbour Deduction Lost Games'
            ws['R5'] = 'Cluster Inference Lost Games'
            ws['S5'] = 'Random Forest Lost Games'
            ws['T5'] = 'Monte Carlo Lost Games'
            ws['Q6'] = '=COUNTIFS(M:M,"Loss",N:N,"Neighbour Deduction")'
            ws['R6'] = '=COUNTIFS(M:M,"Loss",N:N,"Cluster Inference")'
            ws['S6'] = '=COUNTIFS(M:M,"Loss",N:N,"Random Forest")'
            ws['T6'] = '=COUNTIFS(M:M,"Loss",N:N,"Monte Carlo")'
            ws['T7'] = '=COUNTIFS(M:M,"Loss",N:N,"Monte Carlo",D:D,"1")'

            wb.save(self.filename)
            print("Formuły zostały dodane do pliku Excel.")
        except Exception as e:
            print(f"Błąd podczas dodawania formuł do Excela: {e}")
